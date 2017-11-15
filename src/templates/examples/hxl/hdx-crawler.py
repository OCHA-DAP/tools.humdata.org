import ckanapi, json, sys, time
from urllib2 import Request, urlopen, URLError, HTTPError
from itertools import islice
from openpyxl import load_workbook
import hxl
from io import BytesIO
import datetime
import hashlib
import xlrd


DELAY = 2
"""Time delay in seconds between datasets, to give HDX a break."""

CKAN_URL = "https://data.humdata.org"
"""Base URL for the CKAN instance."""

#indexFile = {}

with open('working/index_3350.json') as json_data:
    indexFile = json.load(json_data)


def populateIndex(uniqueTags,sampleData,i,attributes,md5,name,url,package_id):
    includeatts = False
    includefile = False
    print sampleData
    for tag in uniqueTags:
        if tag not in indexFile:
            indexFile[tag] = {'samples':['sample_'+str(i)],'attributes':{},'md5s':[md5]}
            includeatts = True
            includefile = True
        else:
            if md5 not in indexFile[tag]['md5s']:
                includeatts = True
            if len(indexFile[tag]['samples'])<5 and md5 not in indexFile[tag]['md5s']:
                indexFile[tag]['samples'].append('sample_'+str(i))
                indexFile[tag]['md5s'].append(md5)
                includefile = True
    if includeatts == True:
        for key in attributes:
            for att in attributes[key]:
                if att in indexFile[key]['attributes']:
                    indexFile[key]['attributes'][att]+=1
                else:
                    indexFile[key]['attributes'][att]=1
    if includefile == True:
        sample = {'data':sampleData,'name':name,'url':url,'package_id':package_id}
        with open('working/sample_'+str(i)+'.json', 'w') as file:
            json.dump(sample, file)

def processHXLData(dataset):
    x = dataset.values
    if len(dataset.values)>2:
        sample = [dataset.headers,dataset.display_tags,dataset.values[0],dataset.values[1],dataset.values[2]]
        uniqueTags = []
        atts = {}
        tags =''
        for tag in dataset.tags:
            if tag!=None:
                tags +=tag
                if tag not in uniqueTags:
                    uniqueTags.append(tag)
        m = hashlib.md5()
        m.update(tags)
        md5 = m.hexdigest()
        for tag in dataset.display_tags:
            tagAtts = tag.split('+')
            if len(tagAtts)>1:
                if tagAtts[0] not in atts:
                    atts[tagAtts[0]]=[]
                for i in range(1, len(tagAtts)):
                    atts[tagAtts[0]].append(tagAtts[i])
        return [uniqueTags,sample,atts,md5]
    return False

def readCsv(csvLocation):
    try:
        content = urlopen(csvLocation)
    except URLError as e:
        print("CSV Failed to download")
    try:
        print "File downloaded and attempting to read HXL"
        dataset = hxl.data(content).cache()
        output = processHXLData(dataset)
        print "HXL output"
        return output
    except Exception as e:
        print e
        return False

def readXlsx(fileLocation):
    print "Trying to download XLSX"
    try:
        response = urlopen(fileLocation)
        try:
            print "Reading XLSX"
            wb = load_workbook(BytesIO(response.read()))
        except:
            print "Error reading "+ str(fileLocation)
            return False
        sheet = wb.active
        data={}
    except URLError as e:
        print("XLS Failed to download")
    try:
        rows_iter = sheet.iter_rows(min_col=1, min_row=1, max_col=sheet.max_column, max_row=sheet.max_row)
        dataset = [[cell.value for cell in row] for row in rows_iter]
        for i, row in enumerate(dataset):
            for j, cell in enumerate(dataset[i]):
                if isinstance(cell, datetime.date):
                    dataset[i][j] = cell.strftime('%m/%d/%Y')
                elif isinstance(cell, basestring):
                    dataset[i][j] = cell.encode('ascii', 'ignore')
        dataset = hxl.data(dataset).cache()
        output = processHXLData(dataset)
        print "HXL output"
        return output
    except Exception as e:
        print e
        return False

def readXls(fileLocation):
    print "Trying to download XLS"
    try:
        response = urlopen(fileLocation).read()
        try:
            print "Reading XLS"
            wb = xlrd.open_workbook(file_contents=response)
        except Exception as e:
            print e
            print "Error reading "+ str(fileLocation)
            return False
        xl_sheet = wb.sheet_by_index(0)
    except URLError as e:
        print("XLS Failed to download")
    try:
        dataset = []
        for row in range (0, xl_sheet.nrows):
            r = []
            for col in range(0, xl_sheet.ncols):
                if isinstance(xl_sheet.cell_value(row, col), basestring):
                    r.append(xl_sheet.cell_value(row, col).encode('ascii', 'ignore'))
                else:
                    r.append(xl_sheet.cell_value(row, col))
                #if isinstance(cell, datetime.date):
                #    dataset[i][j] = cell.strftime('%m/%d/%Y')
            dataset.append(r)
        dataset = hxl.data(dataset).cache()
        output = processHXLData(dataset)
        print "HXL output"
        return output
    except Exception as e:
        print e
        return False

# find datasets tagged HXL
def find_hxl_datasets(start, rows):
    """Return a page of HXL datasets."""
    return ckan.action.package_search(start=start, rows=rows, fq="tags:hxl")

# Open a connection to HDX
ckan = ckanapi.RemoteCKAN(CKAN_URL)
result_start_pos = 1000
result_page_size = 4000

result = find_hxl_datasets(0, result_page_size)
packages = result["results"]
result2 = find_hxl_datasets(1000, result_page_size)
packages2 = result2["results"]

allpackages = packages + packages2

# Iterate through all the datasets ("packages") and resources on HDX
i=0
for package in allpackages:
    # package = ckan.action.package_show(id=package_id)
    print("Package: " + format(package["title"]))
    print package
    # for each resource in a package (some packages have multiple csv files for example), print the name, url and format
    for resource in package["resources"]:
#        if i>3140:
        if i>3350:
            print "---------------------"
            print("  {}".format(resource["name"].encode('ascii', 'ignore')))
            print("    {}".format(resource["url"]))
            print resource["format"]

            if resource["format"] == "CSV":
                file_data = readCsv(resource["url"])
                if(file_data!=False):
                    populateIndex(file_data[0],file_data[1],i,file_data[2],file_data[3],resource["name"],resource["url"],resource["package_id"])

            if resource["format"] == "XLSX":
                file_data = readXlsx(resource["url"])
                if(file_data!=False):
                    populateIndex(file_data[0],file_data[1],i,file_data[2],file_data[3],resource["name"],resource["url"],resource["package_id"])

            if resource["format"] == "XLS":
                file_data = readXls(resource["url"])
                if(file_data!=False):
                    populateIndex(file_data[0],file_data[1],i,file_data[2],file_data[3],resource["name"],resource["url"],resource["package_id"])

            if i%10==0:
                with open('working/index_'+str(i)+'.json', 'w') as file:
                    json.dump(indexFile, file)
            time.sleep(DELAY) # give HDX a short rest
        i+=1
        print i

